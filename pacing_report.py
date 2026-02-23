"""
pacing_report.py — Pacing Report Generator

Reads a "Billed n Booked"-style sheet from the input workbook and writes a
clean, standalone "Pacing - Yearly Summary" Excel file for FY2026.

Usage:
    python pacing_report.py

Running the script opens file-picker dialogs for the input workbook and the
output save location.  The input file only needs to contain a sheet whose
name starts with "Billed n Booked".

Data flow
---------
BNB sheet rows 94/96/98         → BNB - DAI/Embedded/Social   (all 12 months)
BNB sheet rows 100/102 cols C-E → Podroll / PodcastOne Pro Q4  (Jan-Mar)
FIXED_HYBRID_APR_DEC            → Podroll / PodcastOne Pro Q1-Q3 (Apr-Dec)
FIXED_NON_BNB_MONTHLY           → Programmatic, Art 19, Subscription,
                                   Trade, Other Revenues, Barter (all months)

BNB column mapping (C=Jan … N=Dec, fiscal year Apr-Mar):
  Q1 = Apr+May+Jun = cols F+G+H  (6,7,8)
  Q2 = Jul+Aug+Sep = cols I+J+K  (9,10,11)
  Q3 = Oct+Nov+Dec = cols L+M+N  (12,13,14)
  Q4 = Jan+Feb+Mar = cols C+D+E  (3,4,5)
"""

import argparse
import os
import sys
import tkinter as tk
from tkinter import filedialog

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("openpyxl is required. Install it with: pip install openpyxl")

# ---------------------------------------------------------------------------
# BNB sheet constants
# ---------------------------------------------------------------------------

# Sheet name prefix — matched case-insensitively
BNB_SHEET_PREFIX = "billed n booked"

BNB_COL_LABEL = 2  # Column B — row label

# Fiscal-quarter column groups (1-indexed)
BNB_Q1_COLS = (6, 7, 8)    # Apr, May, Jun
BNB_Q2_COLS = (9, 10, 11)  # Jul, Aug, Sep
BNB_Q3_COLS = (12, 13, 14) # Oct, Nov, Dec
BNB_Q4_COLS = (3, 4, 5)    # Jan, Feb, Mar

# BNB category rows — located by label in col B (no hardcoded row numbers).
# All 12 months (cols C-N) are read directly from the sheet.
# (label in BNB sheet,            label in report)
BNB_ROWS = [
    ("Embedded Billable Total",        "BNB - Embedded Direct"),
    ("DAI Billable Total",             "BNB - DAI Direct"),
    ("Social Billable Total",          "BNB - Other (social/segments)"),
    ("Pod Roll Billable Total",        "Podroll"),
    ("Podcastone Pro Billable Total",  "PodcastOne Pro"),
]

# ---------------------------------------------------------------------------
# Fixed monthly data for non-BNB categories (raw dollars, Apr-Mar order)
# Source: "Pacing - Monthly Inputs (actual" rows 31/32/35/36/37/40, cols C-N
# Order: [Apr, May, Jun, Jul, Aug, Sep, Oct, Nov, Dec, Jan, Feb, Mar]
# Update these values when the source data changes.
# ---------------------------------------------------------------------------

FIXED_NON_BNB_MONTHLY = {
    "Programmatic Revenue": [
        85019.87,  79655.12, 162663.64,   # Q1: Apr, May, Jun
        198940.00, 158421.21, 155048.48,  # Q2: Jul, Aug, Sep
        117238.29, 130928.86, 194080.40,  # Q3: Oct, Nov, Dec
        91169.94,  28025.44,      0.00,   # Q4: Jan, Feb, Mar
    ],
    "Art 19 - MG or billed": [
        509272.01, 536669.03, 798783.13,  # Q1
        656476.97, 709357.47, 839603.10,  # Q2
        593177.86, 727537.02, 520419.54,  # Q3
        96456.84,  43375.69,      0.00,   # Q4
    ],
    "Subscription Revenues": [
        1322.47, 1115.27, 136.38,         # Q1
          21.37,    0.00,   0.00,         # Q2
           0.00,    0.00,   0.00,         # Q3
           0.00,    0.00,   0.00,         # Q4
    ],
    "Trade": [0.00] * 12,
    "Other Revenues": [
        30251.66, 28956.72,  42527.63,    # Q1
        48896.85, 43165.21,  62240.51,    # Q2
        57855.95, 44210.13, 135028.99,    # Q3
            0.00,     0.00,      0.00,    # Q4
    ],
    "Barter": [
        2333333.33, 2333333.33, 2333333.33,  # Q1
        2333333.33, 2333333.33, 2333333.33,  # Q2
        2333333.33, 2333333.33, 2246693.85,  # Q3
        2333333.33, 2333333.33, 2333333.33,  # Q4
    ],
}

# ---------------------------------------------------------------------------
# Fixed FY2025 comparison data — already in $thousands [Q1, Q2, Q3, Q4, FY YTD]
# Source: "Pacing - Monthly Inputs (actual" rows 59-68/71, cols P-T
# These values do not change.
# ---------------------------------------------------------------------------

FIXED_FY2025 = {
    "BNB - Embedded Direct":        [4183.77, 4043.83, 4372.29, 4205.46, 16805.34],
    "BNB - DAI Direct":             [ 633.39,  663.67,  804.12,  646.26,  2747.44],
    "BNB - Other (social/segments)":[   0.00,    0.00,    0.00,  127.15,   127.15],
    "Programmatic Revenue":         [2144.75, 1501.63, 1513.64,  818.18,  5978.20],
    "Art 19 - MG or billed":        [   0.00,    0.00,    0.00,  728.04,   728.04],
    "Podroll":                      [   0.00,    0.00,    0.00,  177.59,   177.59],
    "PodcastOne Pro":               [   0.00,    0.00,    0.00,   65.75,    65.75],
    "Subscription Revenues":        [   5.65,    5.32,    5.45,    4.25,    20.67],
    "Trade":                        [   0.00,    0.00,    0.00,    0.00,     0.00],
    "Other Revenues":               [ 191.12,  -60.91,   14.47,  270.34,   415.02],
    "Barter":                       [6000.00, 6000.00, 6000.00, 7000.00, 25000.00],
}

# Report row order (matches output layout)
ORDERED_LABELS = [
    "BNB - Other (social/segments)",
    "BNB - DAI Direct",
    "BNB - Embedded Direct",
    "Programmatic Revenue",
    "Art 19 - MG or billed",
    "Podroll",
    "PodcastOne Pro",
    "Subscription Revenues",
    "Trade",
    "Other Revenues",
]

# ---------------------------------------------------------------------------
# Output appearance
# ---------------------------------------------------------------------------

OUTPUT_SHEET_NAME = "Pacing - Yearly Summary"
TITLE_TEXT        = "Podcast One — Pacing Report: Yearly Summary"
SUBTITLE_TEXT     = "All figures in $thousands"

COLOR_HEADER_BG = "1F4E79"  # dark blue
COLOR_TOTAL_BG  = "BDD7EE"  # medium blue
COLOR_BARTER_BG = "FFF2CC"  # light yellow
COLOR_WHITE     = "FFFFFF"

NUMBER_FORMAT = '"$"#,##0.00_);("$"#,##0.00)'

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def safe_num(cell) -> float:
    """Return the numeric value of a cell, or 0.0 for None / non-numeric."""
    v = cell.value if hasattr(cell, "value") else cell
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def find_bnb_sheet(wb) -> str:
    """Return the name of the first sheet whose name starts with BNB_SHEET_PREFIX.

    Falls back to the only sheet if the workbook has exactly one sheet.
    Raises SystemExit if no suitable sheet is found.
    """
    for name in wb.sheetnames:
        if name.lower().startswith(BNB_SHEET_PREFIX):
            return name
    if len(wb.sheetnames) == 1:
        print(
            f"WARNING: No sheet starting with '{BNB_SHEET_PREFIX}' found. "
            f"Using the only sheet: '{wb.sheetnames[0]}'",
            file=sys.stderr,
        )
        return wb.sheetnames[0]
    sys.exit(
        f"ERROR: Could not find a sheet whose name starts with "
        f"'{BNB_SHEET_PREFIX}'.\nSheets found: {wb.sheetnames}"
    )


def _warn_label(sheet_name: str, row: int, expected: str, actual) -> None:
    actual_str = str(actual).strip() if actual else "(empty)"
    if actual_str.lower() != expected.lower():
        print(
            f"WARNING [{sheet_name} row {row}]: expected label '{expected}', "
            f"found '{actual_str}'",
            file=sys.stderr,
        )


def _bnb_quarters(ws, row_num: int) -> list:
    """Read a BNB row and return [Q1, Q2, Q3, Q4, FY] in $thousands."""
    q1 = sum(safe_num(ws.cell(row=row_num, column=c)) for c in BNB_Q1_COLS) / 1000.0
    q2 = sum(safe_num(ws.cell(row=row_num, column=c)) for c in BNB_Q2_COLS) / 1000.0
    q3 = sum(safe_num(ws.cell(row=row_num, column=c)) for c in BNB_Q3_COLS) / 1000.0
    q4 = sum(safe_num(ws.cell(row=row_num, column=c)) for c in BNB_Q4_COLS) / 1000.0
    return [q1, q2, q3, q4, q1 + q2 + q3 + q4]


def _fixed_monthly_quarters(monthly_12: list) -> list:
    """Convert [Apr,May,Jun,Jul,Aug,Sep,Oct,Nov,Dec,Jan,Feb,Mar] -> [Q1,Q2,Q3,Q4,FY]."""
    q1 = sum(monthly_12[0:3]) / 1000.0
    q2 = sum(monthly_12[3:6]) / 1000.0
    q3 = sum(monthly_12[6:9]) / 1000.0
    q4 = sum(monthly_12[9:12]) / 1000.0
    return [q1, q2, q3, q4, q1 + q2 + q3 + q4]

# ---------------------------------------------------------------------------
# Data extraction
# ---------------------------------------------------------------------------

def _build_label_index(ws) -> dict:
    """Scan col B of the sheet and return {stripped_lower_label: row_number}."""
    index = {}
    for row in ws.iter_rows(min_col=BNB_COL_LABEL, max_col=BNB_COL_LABEL):
        cell = row[0]
        if cell.value is not None:
            index[str(cell.value).strip().lower()] = cell.row
    return index


def extract_fy2026(ws, sheet_name: str) -> dict:
    """Read FY2026 data from the BNB-style sheet.

    BNB categories are located by scanning column B for their source label,
    then all 12 monthly columns (C-N) are read and aggregated to fiscal quarters.

    Returns {report_label: [Q1, Q2, Q3, Q4, FY_YTD]} in $thousands.
    """
    result = {}
    label_index = _build_label_index(ws)

    for source_label, report_label in BNB_ROWS:
        row_num = label_index.get(source_label.lower())
        if row_num is None:
            print(
                f"WARNING [{sheet_name}]: row '{source_label}' not found — "
                f"{report_label} will be zero.",
                file=sys.stderr,
            )
            result[report_label] = [0.0] * 5
            continue

        q1 = sum(safe_num(ws.cell(row=row_num, column=c)) for c in BNB_Q1_COLS) / 1000.0
        q2 = sum(safe_num(ws.cell(row=row_num, column=c)) for c in BNB_Q2_COLS) / 1000.0
        q3 = sum(safe_num(ws.cell(row=row_num, column=c)) for c in BNB_Q3_COLS) / 1000.0
        q4 = sum(safe_num(ws.cell(row=row_num, column=c)) for c in BNB_Q4_COLS) / 1000.0
        result[report_label] = [q1, q2, q3, q4, q1 + q2 + q3 + q4]

    # Fixed non-BNB rows
    for label, monthly in FIXED_NON_BNB_MONTHLY.items():
        result[label] = _fixed_monthly_quarters(monthly)

    return result

# ---------------------------------------------------------------------------
# Report assembly
# ---------------------------------------------------------------------------

def build_report_data(fy26: dict) -> list[dict]:
    """Assemble ordered rows and compute totals.

    Each entry is one of:
      {"type": "data"|"total"|"barter", "label": str,
       "fy26": [Q1,Q2,Q3,Q4,FY], "fy25": [Q1,Q2,Q3,Q4,FY]}
      {"type": "spacer"}
    """
    rows = []

    for label in ORDERED_LABELS:
        rows.append({
            "type":  "data",
            "label": label,
            "fy26":  fy26.get(label, [0.0] * 5),
            "fy25":  FIXED_FY2025.get(label, [0.0] * 5),
        })

    # Total (Excl Barter) — computed in Python
    totals26 = [sum(r["fy26"][i] for r in rows if r["type"] == "data") for i in range(5)]
    totals25 = [sum(r["fy25"][i] for r in rows if r["type"] == "data") for i in range(5)]
    rows.append({"type": "total", "label": "Total (Excl Barter)",
                 "fy26": totals26, "fy25": totals25})

    rows.append({"type": "spacer"})

    barter26 = fy26.get("Barter", [0.0] * 5)
    barter25 = FIXED_FY2025.get("Barter", [0.0] * 5)
    rows.append({"type": "barter", "label": "Barter",
                 "fy26": barter26, "fy25": barter25})

    incl26 = [totals26[i] + barter26[i] for i in range(5)]
    incl25 = [totals25[i] + barter25[i] for i in range(5)]
    rows.append({"type": "total", "label": "Total (Incl Barter)",
                 "fy26": incl26, "fy25": incl25})

    return rows

# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=None, size=11, italic=False) -> Font:
    kw = {"bold": bold, "size": size, "italic": italic}
    if color:
        kw["color"] = color
    return Font(**kw)

def _border_bottom() -> Border:
    thin = Side(style="thin")
    return Border(bottom=thin)

def _border_top_bottom() -> Border:
    thin = Side(style="thin")
    return Border(top=thin, bottom=thin)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")

def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")

# ---------------------------------------------------------------------------
# Write output
# ---------------------------------------------------------------------------

def write_yearly_summary(report_rows: list[dict], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = OUTPUT_SHEET_NAME

    # Column widths: A=spacer, B=label, C-G=FY2026, H=gap, I-M=FY2025
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    for col in "CDEFG":
        ws.column_dimensions[col].width = 13
    ws.column_dimensions["H"].width = 4
    for col in "IJKLM":
        ws.column_dimensions[col].width = 13

    # ---- Row 2: Title ----
    ws.row_dimensions[2].height = 24
    ws["B2"].value     = TITLE_TEXT
    ws["B2"].font      = _font(bold=True, size=14, color=COLOR_WHITE)
    ws["B2"].fill      = _fill(COLOR_HEADER_BG)
    ws["B2"].alignment = _center()
    ws.merge_cells("B2:M2")

    # ---- Row 3: Subtitle ----
    ws.row_dimensions[3].height = 16
    ws["B3"].value     = SUBTITLE_TEXT
    ws["B3"].font      = _font(italic=True, size=10)
    ws["B3"].alignment = _center()
    ws.merge_cells("B3:M3")

    # ---- Row 4: Fiscal year group labels ----
    ws.row_dimensions[4].height = 18
    ws["C4"].value     = "Fiscal 2026"
    ws["C4"].font      = _font(bold=True, color=COLOR_WHITE)
    ws["C4"].fill      = _fill(COLOR_HEADER_BG)
    ws["C4"].alignment = _center()
    ws.merge_cells("C4:G4")

    ws["I4"].value     = "Fiscal 2025"
    ws["I4"].font      = _font(bold=True, color=COLOR_WHITE)
    ws["I4"].fill      = _fill(COLOR_HEADER_BG)
    ws["I4"].alignment = _center()
    ws.merge_cells("I4:M4")

    # ---- Row 5: Column headers ----
    ws.row_dimensions[5].height = 18
    ws["B5"].value     = "Revenue Category"
    ws["B5"].font      = _font(bold=True, color=COLOR_WHITE)
    ws["B5"].fill      = _fill(COLOR_HEADER_BG)
    ws["B5"].alignment = _left()

    for i, header in enumerate(["Q1", "Q2", "Q3", "Q4", "FY YTD"]):
        for start_col in (3, 9):   # C=3 for FY26, I=9 for FY25
            cell = ws.cell(row=5, column=start_col + i)
            cell.value     = header
            cell.font      = _font(bold=True, color=COLOR_WHITE)
            cell.fill      = _fill(COLOR_HEADER_BG)
            cell.alignment = _center()
            cell.border    = _border_bottom()

    # ---- Data rows (start at row 6) ----
    current_row = 6

    for entry in report_rows:
        row_type = entry["type"]

        if row_type == "spacer":
            ws.row_dimensions[current_row].height = 8
            current_row += 1
            continue

        ws.row_dimensions[current_row].height = 16

        if row_type == "total":
            row_fill   = _fill(COLOR_TOTAL_BG)
            row_font   = _font(bold=True)
            row_border = _border_top_bottom()
        elif row_type == "barter":
            row_fill   = _fill(COLOR_BARTER_BG)
            row_font   = _font()
            row_border = None
        else:
            row_fill   = None
            row_font   = _font()
            row_border = None

        # Label cell (col B)
        lbl = ws.cell(row=current_row, column=2)
        lbl.value     = entry["label"]
        lbl.font      = row_font
        lbl.alignment = _left()
        if row_fill:   lbl.fill   = row_fill
        if row_border: lbl.border = row_border

        # FY2026 data (cols C-G = 3-7) and FY2025 data (cols I-M = 9-13)
        for vals, start_col in ((entry["fy26"], 3), (entry["fy25"], 9)):
            for i, val in enumerate(vals):
                cell = ws.cell(row=current_row, column=start_col + i)
                cell.value         = val
                cell.number_format = NUMBER_FORMAT
                cell.alignment     = _right()
                cell.font          = row_font
                if row_fill:   cell.fill   = row_fill
                if row_border: cell.border = row_border

        current_row += 1

    ws.freeze_panes = "B6"
    wb.save(output_path)
    print(f"Report written to: {output_path}")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    argparse.ArgumentParser(
        description="Generate a Pacing - Yearly Summary Excel report."
    ).parse_args()   # no positional args; keeps --help working

    # Hide the root Tk window — we only want the dialogs
    root = tk.Tk()
    root.withdraw()

    # --- Pick input file ---
    input_file = filedialog.askopenfilename(
        title="Select the Billed n Booked workbook",
        filetypes=[("Excel workbooks", "*.xlsx *.xlsm"), ("All files", "*.*")],
    )
    if not input_file:
        sys.exit("No input file selected — exiting.")

    # --- Pick output location ---
    default_dir  = os.path.dirname(input_file)
    default_name = "Pacing - Yearly Summary.xlsx"
    output_file = filedialog.asksaveasfilename(
        title="Save the Yearly Summary report as...",
        initialdir=default_dir,
        initialfile=default_name,
        defaultextension=".xlsx",
        filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
    )
    if not output_file:
        sys.exit("No output file selected — exiting.")

    print(f"Opening workbook: {input_file}")
    try:
        wb = openpyxl.load_workbook(input_file, data_only=True)
    except Exception as exc:
        sys.exit(f"ERROR: Could not open workbook: {exc}")

    sheet_name = find_bnb_sheet(wb)
    print(f"Using sheet: '{sheet_name}'")
    ws_bnb = wb[sheet_name]

    print("Extracting FY2026 data...")
    fy26 = extract_fy2026(ws_bnb, sheet_name)

    print("Building report...")
    report_rows = build_report_data(fy26)

    # Quick verification summary
    print("\n--- Quick Verification Summary ---")
    for row in report_rows:
        if row["type"] == "spacer":
            continue
        v26 = row["fy26"]
        v25 = row["fy25"]
        print(f"  {row['label']:<35}  "
              f"FY26: Q1={v26[0]:>10.2f}  YTD={v26[4]:>10.2f}  |  "
              f"FY25: Q1={v25[0]:>10.2f}  YTD={v25[4]:>10.2f}")
    print()

    write_yearly_summary(report_rows, output_file)


if __name__ == "__main__":
    main()
